from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Sum
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_http_methods

from apps.business.models import Member

from .models import Item, ItemCategory, ItemMovement


# ── Dashboard ─────────────────────────────────────────────────────────────────

@login_required
@require_http_methods(['GET'])
def inventory_home(request):
    total_items = Item.objects.count()
    total_quantity = Item.objects.aggregate(t=Sum('quantity'))['t'] or 0
    available = Item.objects.filter(status='available').count()
    lent = Item.objects.filter(status='lent').count()
    maintenance = Item.objects.filter(status='maintenance').count()
    lost = Item.objects.filter(status='lost').count()

    categories = ItemCategory.objects.filter(active=True).annotate(items_count=Count('items'))
    recent_movements = ItemMovement.objects.select_related('item', 'member').order_by('-date')[:10]

    context = {
        'total_items': total_items,
        'total_quantity': total_quantity,
        'available': available,
        'lent': lent,
        'maintenance': maintenance,
        'lost': lost,
        'categories': categories,
        'recent_movements': recent_movements,
    }
    return render(request, 'inventory/home.html', context)


# ── Categorías ────────────────────────────────────────────────────────────────

@login_required
@require_http_methods(['GET'])
def categories_list(request):
    categories = ItemCategory.objects.annotate(items_count=Count('items')).order_by('name')
    return render(request, 'inventory/categories/list.html', {'categories': categories})


@login_required
@require_http_methods(['GET', 'POST'])
def category_create(request):
    if request.method == 'POST':
        ItemCategory.objects.create(
            name=request.POST.get('name'),
            description=request.POST.get('description', ''),
            icon=request.POST.get('icon', ''),
        )
        messages.success(request, 'Categoría creada exitosamente.')
        return redirect('inventory:categories_list')
    return render(request, 'inventory/categories/form.html', {'action': 'Crear'})


@login_required
@require_http_methods(['GET', 'POST'])
def category_edit(request, category_id):
    category = get_object_or_404(ItemCategory, id=category_id)
    if request.method == 'POST':
        category.name = request.POST.get('name')
        category.description = request.POST.get('description', '')
        category.icon = request.POST.get('icon', '')
        category.active = request.POST.get('active') == 'on'
        category.save()
        messages.success(request, f'Categoría "{category.name}" actualizada.')
        return redirect('inventory:categories_list')
    return render(request, 'inventory/categories/form.html', {'category': category, 'action': 'Editar'})


@login_required
@require_http_methods(['POST'])
def category_delete(request, category_id):
    category = get_object_or_404(ItemCategory, id=category_id)
    if category.items.exists():
        messages.error(request, f'No se puede eliminar "{category.name}" porque tiene artículos asociados.')
        return redirect('inventory:categories_list')
    category.delete()
    messages.success(request, f'Categoría "{category.name}" eliminada.')
    return redirect('inventory:categories_list')


# ── Artículos ─────────────────────────────────────────────────────────────────

@login_required
@require_http_methods(['GET'])
def items_export(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    items = Item.objects.select_related('category', 'assigned_to').order_by('category__name', 'name')

    # Apply same filters as list view
    search = request.GET.get('q', '').strip()
    if search:
        items = items.filter(Q(name__icontains=search) | Q(code__icontains=search) | Q(description__icontains=search))
    status_filter = request.GET.get('status')
    if status_filter:
        items = items.filter(status=status_filter)
    category_filter = request.GET.get('category')
    if category_filter:
        items = items.filter(category_id=category_filter)
    condition_filter = request.GET.get('condition')
    if condition_filter:
        items = items.filter(condition=condition_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'

    # Header style
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

    headers = [
        'Nombre', 'Categoría', 'Código/Serie', 'Cantidad', 'Estado',
        'Condición', 'Ubicación', 'Responsable', 'Fecha de Compra',
        'Precio', 'Descripción', 'Notas',
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    status_display = dict(Item.STATUS_CHOICES)
    condition_display = dict(Item.CONDITION_CHOICES)

    for row, item in enumerate(items, 2):
        ws.cell(row=row, column=1, value=item.name)
        ws.cell(row=row, column=2, value=item.category.name)
        ws.cell(row=row, column=3, value=item.code)
        ws.cell(row=row, column=4, value=item.quantity)
        ws.cell(row=row, column=5, value=status_display.get(item.status, item.status))
        ws.cell(row=row, column=6, value=condition_display.get(item.condition, item.condition))
        ws.cell(row=row, column=7, value=item.location)
        ws.cell(row=row, column=8, value=item.assigned_to.get_full_name() if item.assigned_to else '')
        ws.cell(row=row, column=9, value=item.purchase_date.strftime('%d/%m/%Y') if item.purchase_date else '')
        ws.cell(row=row, column=10, value=float(item.purchase_price) if item.purchase_price else '')
        ws.cell(row=row, column=11, value=item.description)
        ws.cell(row=row, column=12, value=item.notes)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario.xlsx"'
    wb.save(response)
    return response


@login_required
@require_http_methods(['GET'])
def items_list(request):
    items = Item.objects.select_related('category', 'assigned_to').order_by('category__name', 'name')

    search = request.GET.get('q', '').strip()
    if search:
        items = items.filter(Q(name__icontains=search) | Q(code__icontains=search) | Q(description__icontains=search))

    status_filter = request.GET.get('status')
    if status_filter:
        items = items.filter(status=status_filter)

    category_filter = request.GET.get('category')
    if category_filter:
        items = items.filter(category_id=category_filter)

    condition_filter = request.GET.get('condition')
    if condition_filter:
        items = items.filter(condition=condition_filter)

    context = {
        'items': items,
        'categories': ItemCategory.objects.filter(active=True).order_by('name'),
        'search': search,
        'status_filter': status_filter,
        'category_filter': category_filter,
        'condition_filter': condition_filter,
    }
    return render(request, 'inventory/items/list.html', context)


@login_required
@require_http_methods(['GET', 'POST'])
def item_create(request):
    if request.method == 'POST':
        member_id = request.POST.get('assigned_to')
        Item.objects.create(
            name=request.POST.get('name'),
            description=request.POST.get('description', ''),
            category_id=request.POST.get('category'),
            code=request.POST.get('code', ''),
            quantity=request.POST.get('quantity', 1),
            condition=request.POST.get('condition', 'good'),
            status=request.POST.get('status', 'available'),
            location=request.POST.get('location', ''),
            purchase_date=request.POST.get('purchase_date') or None,
            purchase_price=request.POST.get('purchase_price') or None,
            assigned_to_id=member_id if member_id else None,
            notes=request.POST.get('notes', ''),
        )
        messages.success(request, 'Artículo registrado exitosamente.')
        return redirect('inventory:items_list')

    context = {
        'categories': ItemCategory.objects.filter(active=True).order_by('name'),
        'members': Member.objects.filter(status='active').order_by('last_name', 'first_name'),
        'action': 'Registrar',
    }
    return render(request, 'inventory/items/form.html', context)


@login_required
@require_http_methods(['GET'])
def item_detail(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    movements = item.movements.select_related('member').order_by('-date')[:20]
    members = Member.objects.filter(status='active').order_by('last_name', 'first_name')
    context = {'item': item, 'movements': movements, 'members': members}
    return render(request, 'inventory/items/detail.html', context)


@login_required
@require_http_methods(['GET', 'POST'])
def item_edit(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    if request.method == 'POST':
        member_id = request.POST.get('assigned_to')
        item.name = request.POST.get('name')
        item.description = request.POST.get('description', '')
        item.category_id = request.POST.get('category')
        item.code = request.POST.get('code', '')
        item.quantity = request.POST.get('quantity', 1)
        item.condition = request.POST.get('condition', 'good')
        item.status = request.POST.get('status', 'available')
        item.location = request.POST.get('location', '')
        item.purchase_date = request.POST.get('purchase_date') or None
        item.purchase_price = request.POST.get('purchase_price') or None
        item.assigned_to_id = member_id if member_id else None
        item.notes = request.POST.get('notes', '')
        item.save()
        messages.success(request, f'Artículo "{item.name}" actualizado.')
        return redirect('inventory:item_detail', item_id=item.id)

    context = {
        'item': item,
        'categories': ItemCategory.objects.filter(active=True).order_by('name'),
        'members': Member.objects.filter(status='active').order_by('last_name', 'first_name'),
        'action': 'Editar',
    }
    return render(request, 'inventory/items/form.html', context)


# ── Movimientos ───────────────────────────────────────────────────────────────

@login_required
@require_http_methods(['POST'])
def item_add_movement(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    movement_type = request.POST.get('movement_type')
    member_id = request.POST.get('member')

    ItemMovement.objects.create(
        item=item,
        movement_type=movement_type,
        member_id=member_id if member_id else None,
        expected_return_date=request.POST.get('expected_return_date') or None,
        notes=request.POST.get('notes', ''),
    )

    # Update item status based on movement
    status_map = {
        'checkout': 'lent',
        'return': 'available',
        'maintenance': 'maintenance',
        'maintenance_done': 'available',
        'lost': 'lost',
        'found': 'available',
        'retired': 'retired',
    }
    new_status = status_map.get(movement_type)
    if new_status:
        item.status = new_status
        # Update assigned_to on checkout/return
        if movement_type == 'checkout' and member_id:
            item.assigned_to_id = member_id
        elif movement_type in ('return', 'maintenance_done', 'found'):
            item.assigned_to = None
        item.save()

    messages.success(request, f'Movimiento registrado: {item.get_status_display()}')
    return redirect('inventory:item_detail', item_id=item.id)
