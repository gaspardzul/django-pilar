import secrets

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import models
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_http_methods

from apps.business.models import (
    Member, Ministry, Family, MemberMinistry, FamilyMember,
    Event, EventWorkGroup, EventWorker,
    EventBudget, EventIncome, EventExpense,
    EventLodging, LodgingHost, LodgingGuest,
)
from .models import SubscriptionPlan, UserSettings
from .tasks import (
    send_subscription_cancellation_email,
    send_subscription_confirmation_email,
    send_trial_started_email,
)


@login_required
@require_http_methods(['GET'])
def dashboard_home(request):
    # Get statistics
    total_members = Member.objects.filter(status='active').count()
    total_ministries = Ministry.objects.filter(active=True).count()
    total_families = Family.objects.count()
    
    # Recent members (last 5)
    recent_members = Member.objects.all().order_by('-created_at')[:5]
    
    context = {
        'total_members': total_members,
        'total_ministries': total_ministries,
        'total_families': total_families,
        'recent_members': recent_members,
    }
    return render(request, 'dashboard/home.html', context)

@login_required
@require_http_methods(['GET', 'POST'])
def profile(request):
    if request.method == 'POST':
        # Handle profile update
        user = request.user
        user.first_name = request.POST.get('first_name', '')
        user.last_name = request.POST.get('last_name', '')
        user.save()
        messages.success(request, 'Profile updated successfully.')
        return redirect('dashboard:profile')
    return render(request, 'dashboard/profile.html')

@login_required
@require_http_methods(['GET', 'POST'])
def settings(request):
    user_settings, created = UserSettings.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        user_settings.notify_comments = request.POST.get('comments') == 'on'
        user_settings.notify_updates = request.POST.get('updates') == 'on'
        user_settings.notify_marketing = request.POST.get('marketing') == 'on'
        user_settings.save()

        messages.success(request, 'Settings updated successfully.')
        return redirect('dashboard:settings')

    # Check if a new API key was just generated (stored in session)
    new_api_key = request.session.pop('new_api_key', None)

    context = {
        'notification_settings': {
            'comments': user_settings.notify_comments,
            'updates': user_settings.notify_updates,
            'marketing': user_settings.notify_marketing,
        },
        'subscription': {
            'plan': user_settings.subscription_plan,
            'plan_name': user_settings.subscription_plan.name if user_settings.subscription_plan else None,
            'status': user_settings.subscription_status,
            'is_active': user_settings.is_subscription_active,
            'is_trial': user_settings.is_trial_active,
            'start_date': user_settings.subscription_start_date,
            'end_date': user_settings.subscription_end_date,
            'trial_end_date': user_settings.trial_end_date,
        },
        'api': {
            'has_key': bool(user_settings.api_key),
            'key_created_at': user_settings.api_key_created_at,
            'new_key': new_api_key,
        },
    }
    return render(request, 'dashboard/settings.html', context)

@login_required
@require_http_methods(['POST'])
def generate_api_key(request):
    user_settings, created = UserSettings.objects.get_or_create(user=request.user)

    api_key = secrets.token_urlsafe(32)
    user_settings.api_key = api_key
    user_settings.api_key_created_at = timezone.now()
    user_settings.save()

    # Store the key in session so it can be shown once on the settings page
    request.session['new_api_key'] = api_key

    messages.success(request, 'API key generated. Copy it now — it won\'t be shown again.')
    return redirect('dashboard:settings')

@login_required
@require_http_methods(['GET'])
def subscription_plans(request):
    plans = SubscriptionPlan.objects.filter(is_active=True)
    user_settings, created = UserSettings.objects.get_or_create(user=request.user)

    context = {
        'plans': plans,
        'current_plan': user_settings.subscription_plan,
        'subscription_status': user_settings.subscription_status,
        'is_subscription_active': user_settings.is_subscription_active,
        'is_trial_active': user_settings.is_trial_active,
    }
    return render(request, 'dashboard/subscription_plans.html', context)

@login_required
@require_http_methods(['POST'])
def subscribe_to_plan(request, plan_slug):
    plan = get_object_or_404(SubscriptionPlan, slug=plan_slug, is_active=True)
    user_settings = UserSettings.objects.get(user=request.user)

    # Check if user already has an active subscription
    if user_settings.is_subscription_active:
        messages.warning(request, 'You already have an active subscription.')
        return redirect('dashboard:subscription_plans')

    # Update user settings with new subscription
    user_settings.subscription_plan = plan
    user_settings.subscription_status = 'active'
    user_settings.subscription_start_date = timezone.now()

    # Set subscription end date based on interval
    if plan.interval == 'monthly':
        user_settings.subscription_end_date = timezone.now() + timezone.timedelta(days=30)
    else:  # yearly
        user_settings.subscription_end_date = timezone.now() + timezone.timedelta(days=365)

    user_settings.save()

    send_subscription_confirmation_email.enqueue(
        user_email=request.user.email,
        plan_name=plan.name,
    )

    messages.success(request, f'Successfully subscribed to {plan.name} plan.')
    return redirect('dashboard:settings')

@login_required
@require_http_methods(['POST'])
def cancel_subscription(request):
    user_settings = UserSettings.objects.get(user=request.user)

    if not user_settings.is_subscription_active:
        messages.warning(request, 'You do not have an active subscription to cancel.')
        return redirect('dashboard:settings')

    user_settings.subscription_status = 'cancelled'
    user_settings.save()

    send_subscription_cancellation_email.enqueue(user_email=request.user.email)

    messages.success(request, 'Your subscription has been cancelled.')
    return redirect('dashboard:settings')

@login_required
@require_http_methods(['POST'])
def start_trial(request):
    user_settings = UserSettings.objects.get(user=request.user)

    if user_settings.is_subscription_active or user_settings.is_trial_active:
        messages.warning(request, 'You already have an active subscription or trial.')
        return redirect('dashboard:subscription_plans')

    # Start trial period (14 days)
    user_settings.subscription_status = 'trial'
    user_settings.trial_end_date = timezone.now() + timezone.timedelta(days=14)
    user_settings.save()

    send_trial_started_email.enqueue(user_email=request.user.email)

    messages.success(request, 'Trial period started successfully.')
    return redirect('dashboard:settings')


# ============================================
# CHURCH MANAGEMENT VIEWS
# ============================================

# Members Views
@login_required
@require_http_methods(['GET'])
def members_list(request):
    members = Member.objects.all().select_related().prefetch_related('ministry_memberships__ministry')
    
    # Search query
    search_query = request.GET.get('q', '').strip()
    if search_query:
        members = members.filter(
            models.Q(first_name__icontains=search_query) |
            models.Q(last_name__icontains=search_query) |
            models.Q(email__icontains=search_query) |
            models.Q(phone__icontains=search_query)
        )
    
    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter:
        members = members.filter(status=status_filter)
    
    # Filter by gender
    gender_filter = request.GET.get('gender')
    if gender_filter:
        members = members.filter(gender=gender_filter)
    
    # Filter by baptism status
    baptism_filter = request.GET.get('baptism')
    if baptism_filter == 'baptized':
        members = members.filter(is_baptized=True)
    elif baptism_filter == 'not_baptized':
        members = members.filter(is_baptized=False)
    
    # Filter by ministry
    ministry_filter = request.GET.get('ministry')
    if ministry_filter:
        members = members.filter(ministry_memberships__ministry_id=ministry_filter, ministry_memberships__end_date__isnull=True).distinct()
    
    # Filter by age range
    age_min = request.GET.get('age_min')
    age_max = request.GET.get('age_max')
    if age_min or age_max:
        from datetime import date
        today = date.today()
        if age_max:
            # Max age means born after this date
            date_max = date(today.year - int(age_max), today.month, today.day)
            members = members.filter(date_of_birth__gte=date_max)
        if age_min:
            # Min age means born before this date
            date_min = date(today.year - int(age_min), today.month, today.day)
            members = members.filter(date_of_birth__lte=date_min)
    
    # Order by
    order_by = request.GET.get('order_by', 'name')
    if order_by == 'name':
        members = members.order_by('last_name', 'first_name')
    elif order_by == 'join_date':
        members = members.order_by('-join_date')
    elif order_by == 'baptism_date':
        members = members.order_by('-baptism_date')
    
    # Get all ministries for filter dropdown
    ministries = Ministry.objects.filter(active=True).order_by('name')
    
    context = {
        'members': members,
        'total_members': Member.objects.filter(status='active').count(),
        'ministries': ministries,
        'search_query': search_query,
        'status_filter': status_filter,
        'gender_filter': gender_filter,
        'baptism_filter': baptism_filter,
        'ministry_filter': ministry_filter,
        'age_min': age_min,
        'age_max': age_max,
        'order_by': order_by,
    }
    return render(request, 'dashboard/members/list.html', context)


@login_required
@require_http_methods(['GET'])
def member_detail(request, member_id):
    member = get_object_or_404(Member, id=member_id)
    
    # Get ministries
    ministries = member.ministry_memberships.filter(end_date__isnull=True)
    
    # Get family relationships
    families = member.family_relationships.filter(end_date__isnull=True)
    
    # Get available ministries (not already in)
    current_ministry_ids = ministries.values_list('ministry_id', flat=True)
    available_ministries = Ministry.objects.filter(active=True).exclude(id__in=current_ministry_ids).order_by('name')
    
    # Get available families (not already in)
    current_family_ids = families.values_list('family_id', flat=True)
    available_families = Family.objects.exclude(id__in=current_family_ids).order_by('family_name')
    
    context = {
        'member': member,
        'ministries': ministries,
        'families': families,
        'available_ministries': available_ministries,
        'available_families': available_families,
    }
    return render(request, 'dashboard/members/detail.html', context)


@login_required
@require_http_methods(['GET', 'POST'])
def member_create(request):
    if request.method == 'POST':
        member = Member.objects.create(
            first_name=request.POST.get('first_name'),
            last_name=request.POST.get('last_name'),
            date_of_birth=request.POST.get('date_of_birth') or None,
            gender=request.POST.get('gender', ''),
            email=request.POST.get('email', ''),
            phone=request.POST.get('phone', ''),
            address=request.POST.get('address', ''),
            join_date=request.POST.get('join_date') or timezone.now().date(),
            status=request.POST.get('status', 'active'),
            notes=request.POST.get('notes', ''),
            is_baptized=request.POST.get('is_baptized') == 'true',
            baptism_date=request.POST.get('baptism_date') or None,
            baptism_place=request.POST.get('baptism_place', ''),
            baptized_by=request.POST.get('baptized_by', ''),
        )
        messages.success(request, f'Miembro {member.get_full_name()} creado exitosamente.')
        
        # Check which action was clicked
        action = request.POST.get('action', 'save_and_return')
        if action == 'save_and_continue':
            return redirect('dashboard:member_edit', member_id=member.id)
        else:
            return redirect('dashboard:members_list')
    
    return render(request, 'dashboard/members/form.html', {'action': 'Create'})


@login_required
@require_http_methods(['GET', 'POST'])
def member_edit(request, member_id):
    member = get_object_or_404(Member, id=member_id)
    
    if request.method == 'POST':
        member.first_name = request.POST.get('first_name')
        member.last_name = request.POST.get('last_name')
        member.date_of_birth = request.POST.get('date_of_birth') or None
        member.gender = request.POST.get('gender', '')
        member.email = request.POST.get('email', '')
        member.phone = request.POST.get('phone', '')
        member.address = request.POST.get('address', '')
        member.join_date = request.POST.get('join_date') or timezone.now().date()
        member.status = request.POST.get('status', 'active')
        member.notes = request.POST.get('notes', '')
        member.is_baptized = request.POST.get('is_baptized') == 'true'
        member.baptism_date = request.POST.get('baptism_date') or None
        member.baptism_place = request.POST.get('baptism_place', '')
        member.baptized_by = request.POST.get('baptized_by', '')
        member.save()
        
        messages.success(request, f'Miembro {member.get_full_name()} actualizado exitosamente.')
        
        # Check which action was clicked
        action = request.POST.get('action', 'save_and_return')
        if action == 'save_and_continue':
            return redirect('dashboard:member_edit', member_id=member.id)
        else:
            return redirect('dashboard:members_list')
    
    context = {'member': member, 'action': 'Edit'}
    return render(request, 'dashboard/members/form.html', context)


@login_required
@require_http_methods(['POST'])
def member_add_to_ministry(request, member_id):
    member = get_object_or_404(Member, id=member_id)
    ministry_id = request.POST.get('ministry_id')
    
    if ministry_id:
        ministry = get_object_or_404(Ministry, id=ministry_id)
        
        # Check if member is already in this ministry
        existing = MemberMinistry.objects.filter(
            ministry=ministry,
            member=member,
            end_date__isnull=True
        ).exists()
        
        if existing:
            messages.warning(request, f'Ya estás en el ministerio {ministry.name}.')
        else:
            MemberMinistry.objects.create(
                ministry=ministry,
                member=member,
                role=request.POST.get('role', 'member'),
                start_date=request.POST.get('start_date') or None,
                notes=request.POST.get('notes', ''),
            )
            messages.success(request, f'Agregado al ministerio {ministry.name} exitosamente.')
    
    return redirect('dashboard:member_detail', member_id=member_id)


@login_required
@require_http_methods(['POST'])
def member_add_to_family(request, member_id):
    member = get_object_or_404(Member, id=member_id)
    family_id = request.POST.get('family_id')
    
    if family_id:
        family = get_object_or_404(Family, id=family_id)
        
        # Check if member is already in this family
        existing = FamilyMember.objects.filter(
            family=family,
            member=member,
            end_date__isnull=True
        ).exists()
        
        if existing:
            messages.warning(request, f'Ya estás en la familia {family.family_name}.')
        else:
            family_member = FamilyMember.objects.create(
                family=family,
                member=member,
                relationship_type=request.POST.get('relationship_type'),
                is_primary_contact=request.POST.get('is_primary_contact') == 'on',
                notes=request.POST.get('notes', ''),
            )
            
            # Update family primary contact if requested
            if family_member.is_primary_contact:
                family.primary_contact = member
                family.save()
            
            messages.success(request, f'Agregado a la familia {family.family_name} exitosamente.')
    
    return redirect('dashboard:member_detail', member_id=member_id)


# Ministries Views
@login_required
@require_http_methods(['GET'])
def ministries_list(request):
    ministries = Ministry.objects.all().order_by('name')
    
    # Search
    search_query = request.GET.get('q')
    if search_query:
        ministries = ministries.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Filter by active status
    active_filter = request.GET.get('active')
    if active_filter == 'true':
        ministries = ministries.filter(active=True)
    elif active_filter == 'false':
        ministries = ministries.filter(active=False)
    
    context = {
        'ministries': ministries,
        'total_ministries': Ministry.objects.filter(active=True).count(),
        'active_filter': active_filter,
        'search_query': search_query,
    }
    return render(request, 'dashboard/ministries/list.html', context)


@login_required
@require_http_methods(['GET'])
def ministry_detail(request, ministry_id):
    ministry = get_object_or_404(Ministry, id=ministry_id)
    
    # Get active members
    active_members = ministry.get_active_members()
    
    # Get leaders count
    leaders_count = active_members.filter(role__in=['leader', 'co_leader']).count()
    
    # Get available members (not already in this ministry)
    current_member_ids = active_members.values_list('member_id', flat=True)
    available_members = Member.objects.filter(status='active').exclude(id__in=current_member_ids).order_by('last_name', 'first_name')
    
    context = {
        'ministry': ministry,
        'active_members': active_members,
        'leaders_count': leaders_count,
        'available_members': available_members,
    }
    return render(request, 'dashboard/ministries/detail.html', context)


@login_required
@require_http_methods(['GET', 'POST'])
def ministry_create(request):
    if request.method == 'POST':
        ministry = Ministry.objects.create(
            name=request.POST.get('name'),
            description=request.POST.get('description', ''),
            active=request.POST.get('active') == 'on',
        )
        
        # Set leader if provided
        leader_id = request.POST.get('leader')
        if leader_id:
            ministry.leader = Member.objects.get(id=leader_id)
            ministry.save()
        
        messages.success(request, f'Ministry {ministry.name} created successfully.')
        return redirect('dashboard:ministry_detail', ministry_id=ministry.id)
    
    members = Member.objects.filter(status='active').order_by('last_name', 'first_name')
    return render(request, 'dashboard/ministries/form.html', {'action': 'Create', 'members': members})


@login_required
@require_http_methods(['GET', 'POST'])
def ministry_edit(request, ministry_id):
    ministry = get_object_or_404(Ministry, id=ministry_id)
    
    if request.method == 'POST':
        ministry.name = request.POST.get('name')
        ministry.description = request.POST.get('description', '')
        ministry.active = request.POST.get('active') == 'on'
        
        # Update leader if provided
        leader_id = request.POST.get('leader')
        if leader_id:
            ministry.leader = Member.objects.get(id=leader_id)
        else:
            ministry.leader = None
        
        ministry.save()
        
        messages.success(request, f'Ministry {ministry.name} updated successfully.')
        return redirect('dashboard:ministry_detail', ministry_id=ministry.id)
    
    members = Member.objects.filter(status='active').order_by('last_name', 'first_name')
    context = {'ministry': ministry, 'action': 'Edit', 'members': members}
    return render(request, 'dashboard/ministries/form.html', context)


@login_required
@require_http_methods(['POST'])
def ministry_add_member(request, ministry_id):
    ministry = get_object_or_404(Ministry, id=ministry_id)
    member_id = request.POST.get('member_id')
    
    if member_id:
        member = get_object_or_404(Member, id=member_id)
        
        # Check if member is already in this ministry
        existing = MemberMinistry.objects.filter(
            ministry=ministry,
            member=member,
            end_date__isnull=True
        ).exists()
        
        if existing:
            messages.warning(request, f'{member.get_full_name()} ya está en este ministerio.')
        else:
            MemberMinistry.objects.create(
                ministry=ministry,
                member=member,
                role=request.POST.get('role', 'member'),
                start_date=request.POST.get('start_date') or None,
                notes=request.POST.get('notes', ''),
            )
            messages.success(request, f'{member.get_full_name()} agregado al ministerio exitosamente.')
    
    return redirect('dashboard:ministry_detail', ministry_id=ministry_id)


@login_required
@require_http_methods(['POST'])
def ministry_remove_member(request, ministry_id, member_ministry_id):
    ministry = get_object_or_404(Ministry, id=ministry_id)
    member_ministry = get_object_or_404(MemberMinistry, id=member_ministry_id, ministry=ministry)
    
    member_name = member_ministry.member.get_full_name()
    member_ministry.delete()
    
    messages.success(request, f'{member_name} removido del ministerio.')
    return redirect('dashboard:ministry_detail', ministry_id=ministry_id)


# Families Views
@login_required
@require_http_methods(['GET'])
def families_list(request):
    families = Family.objects.all().order_by('family_name')
    
    # Search
    search_query = request.GET.get('q')
    if search_query:
        families = families.filter(family_name__icontains=search_query)
    
    context = {
        'families': families,
        'total_families': Family.objects.count(),
        'search_query': search_query,
    }
    return render(request, 'dashboard/families/list.html', context)


@login_required
@require_http_methods(['GET'])
def family_detail(request, family_id):
    family = get_object_or_404(Family, id=family_id)
    
    # Get family structure
    structure = family.get_family_structure()
    all_members = family.get_family_members()
    
    # Get available members (not already in this family)
    current_member_ids = all_members.values_list('member_id', flat=True)
    available_members = Member.objects.filter(status='active').exclude(id__in=current_member_ids).order_by('last_name', 'first_name')
    
    context = {
        'family': family,
        'structure': structure,
        'all_members': all_members,
        'available_members': available_members,
    }
    return render(request, 'dashboard/families/detail.html', context)


@login_required
@require_http_methods(['GET', 'POST'])
def family_create(request):
    if request.method == 'POST':
        family = Family.objects.create(
            family_name=request.POST.get('family_name'),
            address=request.POST.get('address', ''),
            notes=request.POST.get('notes', ''),
        )
        
        # Set primary contact if provided
        primary_contact_id = request.POST.get('primary_contact')
        if primary_contact_id:
            family.primary_contact = Member.objects.get(id=primary_contact_id)
            family.save()
        
        messages.success(request, f'Family {family.family_name} created successfully.')
        return redirect('dashboard:family_detail', family_id=family.id)
    
    members = Member.objects.filter(status='active').order_by('last_name', 'first_name')
    return render(request, 'dashboard/families/form.html', {'action': 'Create', 'members': members})


@login_required
@require_http_methods(['GET', 'POST'])
def family_edit(request, family_id):
    family = get_object_or_404(Family, id=family_id)
    
    if request.method == 'POST':
        family.family_name = request.POST.get('family_name')
        family.address = request.POST.get('address', '')
        family.notes = request.POST.get('notes', '')
        
        # Update primary contact if provided
        primary_contact_id = request.POST.get('primary_contact')
        if primary_contact_id:
            family.primary_contact = Member.objects.get(id=primary_contact_id)
        else:
            family.primary_contact = None
        
        family.save()
        
        messages.success(request, f'Family {family.family_name} updated successfully.')
        return redirect('dashboard:family_detail', family_id=family.id)
    
    members = Member.objects.filter(status='active').order_by('last_name', 'first_name')
    context = {'family': family, 'action': 'Edit', 'members': members}
    return render(request, 'dashboard/families/form.html', context)


@login_required
@require_http_methods(['POST'])
def family_add_member(request, family_id):
    family = get_object_or_404(Family, id=family_id)
    member_id = request.POST.get('member_id')
    
    if member_id:
        member = get_object_or_404(Member, id=member_id)
        
        # Check if member is already in this family
        existing = FamilyMember.objects.filter(
            family=family,
            member=member,
            end_date__isnull=True
        ).exists()
        
        if existing:
            messages.warning(request, f'{member.get_full_name()} ya está en esta familia.')
        else:
            family_member = FamilyMember.objects.create(
                family=family,
                member=member,
                relationship_type=request.POST.get('relationship_type'),
                is_primary_contact=request.POST.get('is_primary_contact') == 'on',
                notes=request.POST.get('notes', ''),
            )
            
            # Update family primary contact if requested
            if family_member.is_primary_contact:
                family.primary_contact = member
                family.save()
            
            messages.success(request, f'{member.get_full_name()} agregado a la familia exitosamente.')
    
    return redirect('dashboard:family_detail', family_id=family_id)


@login_required
@require_http_methods(['POST'])
def family_remove_member(request, family_id, family_member_id):
    family = get_object_or_404(Family, id=family_id)
    family_member = get_object_or_404(FamilyMember, id=family_member_id, family=family)
    
    member_name = family_member.member.get_full_name()
    family_member.delete()
    
    messages.success(request, f'{member_name} removido de la familia.')
    return redirect('dashboard:family_detail', family_id=family_id)


# Events Views
@login_required
@require_http_methods(['GET'])
def events_list(request):
    events = Event.objects.prefetch_related('lodging').order_by('-start_date')

    search_query = request.GET.get('q')
    if search_query:
        events = events.filter(
            models.Q(name__icontains=search_query) |
            models.Q(description__icontains=search_query) |
            models.Q(location__icontains=search_query)
        )

    status_filter = request.GET.get('status')
    if status_filter:
        events = events.filter(status=status_filter)

    type_filter = request.GET.get('type')
    if type_filter:
        events = events.filter(event_type=type_filter)

    view_mode = request.GET.get('view', 'card')
    if view_mode not in ('card', 'table'):
        view_mode = 'card'

    all_events = Event.objects.all()
    events_with_lodging = all_events.filter(requires_lodging=True).count()
    # Pending = requires lodging but capacity < total_needed
    lodging_pending = sum(
        1 for e in all_events.filter(requires_lodging=True).prefetch_related('lodging')
        if hasattr(e, 'lodging') and not e.lodging.is_covered()
    )

    context = {
        'events': events,
        'total_events': all_events.count(),
        'status_filter': status_filter,
        'type_filter': type_filter,
        'search_query': search_query,
        'view_mode': view_mode,
        'events_with_lodging': events_with_lodging,
        'lodging_pending': lodging_pending,
    }
    return render(request, 'dashboard/events/list.html', context)


@login_required
@require_http_methods(['GET'])
def event_detail(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    
    # Get work groups with their workers
    work_groups = event.work_groups.all()
    
    # Get available members for adding to work groups
    available_members = Member.objects.filter(status='active').order_by('last_name', 'first_name')
    
    # Get or create budget
    budget, created = EventBudget.objects.get_or_create(event=event)
    
    # Get incomes and expenses
    incomes = budget.incomes.all()
    expenses = budget.expenses.all()
    
    context = {
        'event': event,
        'work_groups': work_groups,
        'available_members': available_members,
        'budget': budget,
        'incomes': incomes,
        'expenses': expenses,
    }
    return render(request, 'dashboard/events/detail.html', context)


@login_required
@require_http_methods(['GET', 'POST'])
def event_create(request):
    if request.method == 'POST':
        event = Event.objects.create(
            name=request.POST.get('name'),
            description=request.POST.get('description', ''),
            location=request.POST.get('location'),
            start_date=request.POST.get('start_date'),
            end_date=request.POST.get('end_date') or None,
            is_free_entry=request.POST.get('is_free_entry') == 'on',
            ticket_price=request.POST.get('ticket_price') or None,
            max_capacity=request.POST.get('max_capacity') or None,
            event_type=request.POST.get('event_type', 'service'),
            status=request.POST.get('status', 'draft'),
            notes=request.POST.get('notes', ''),
            requires_lodging=request.POST.get('requires_lodging') == 'on',
        )
        
        # Set organizer if provided
        organizer_id = request.POST.get('organizer')
        if organizer_id:
            event.organizer = Member.objects.get(id=organizer_id)
            event.save()
        
        messages.success(request, f'Evento {event.name} creado exitosamente.')
        return redirect('dashboard:event_detail', event_id=event.id)
    
    members = Member.objects.filter(status='active').order_by('last_name', 'first_name')
    return render(request, 'dashboard/events/form.html', {'action': 'Crear', 'members': members})


@login_required
@require_http_methods(['GET', 'POST'])
def event_edit(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    
    if request.method == 'POST':
        event.name = request.POST.get('name')
        event.description = request.POST.get('description', '')
        event.location = request.POST.get('location')
        event.start_date = request.POST.get('start_date')
        event.end_date = request.POST.get('end_date') or None
        event.is_free_entry = request.POST.get('is_free_entry') == 'on'
        event.ticket_price = request.POST.get('ticket_price') or None
        event.max_capacity = request.POST.get('max_capacity') or None
        event.event_type = request.POST.get('event_type', 'service')
        event.status = request.POST.get('status', 'draft')
        event.notes = request.POST.get('notes', '')
        event.requires_lodging = request.POST.get('requires_lodging') == 'on'
        
        # Update organizer if provided
        organizer_id = request.POST.get('organizer')
        if organizer_id:
            event.organizer = Member.objects.get(id=organizer_id)
        else:
            event.organizer = None
        
        event.save()
        
        messages.success(request, f'Evento {event.name} actualizado exitosamente.')
        return redirect('dashboard:event_detail', event_id=event.id)
    
    members = Member.objects.filter(status='active').order_by('last_name', 'first_name')
    context = {'event': event, 'action': 'Editar', 'members': members}
    return render(request, 'dashboard/events/form.html', context)


# Event Work Groups Views
@login_required
@require_http_methods(['POST'])
def event_add_work_group(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    
    work_group = EventWorkGroup.objects.create(
        event=event,
        name=request.POST.get('name'),
        description=request.POST.get('description', ''),
        group_type=request.POST.get('group_type'),
        required_workers=request.POST.get('required_workers', 1),
        notes=request.POST.get('notes', ''),
    )
    
    # Set coordinator if provided
    coordinator_id = request.POST.get('coordinator')
    if coordinator_id:
        work_group.coordinator = Member.objects.get(id=coordinator_id)
        work_group.save()
    
    messages.success(request, f'Grupo de trabajo {work_group.name} agregado exitosamente.')
    return redirect('dashboard:event_detail', event_id=event_id)


@login_required
@require_http_methods(['POST'])
def event_remove_work_group(request, event_id, work_group_id):
    event = get_object_or_404(Event, id=event_id)
    work_group = get_object_or_404(EventWorkGroup, id=work_group_id, event=event)
    
    group_name = work_group.name
    work_group.delete()
    
    messages.success(request, f'Grupo de trabajo {group_name} eliminado.')
    return redirect('dashboard:event_detail', event_id=event_id)


@login_required
@require_http_methods(['POST'])
def event_add_worker(request, event_id, work_group_id):
    event = get_object_or_404(Event, id=event_id)
    work_group = get_object_or_404(EventWorkGroup, id=work_group_id, event=event)
    member_id = request.POST.get('member_id')
    
    if member_id:
        member = get_object_or_404(Member, id=member_id)
        
        # Check if worker already exists
        existing = EventWorker.objects.filter(
            work_group=work_group,
            member=member
        ).exists()
        
        if existing:
            messages.warning(request, f'{member.get_full_name()} ya está en este grupo de trabajo.')
        else:
            EventWorker.objects.create(
                work_group=work_group,
                member=member,
                status=request.POST.get('status', 'invited'),
                role=request.POST.get('role', ''),
                notes=request.POST.get('notes', ''),
            )
            messages.success(request, f'{member.get_full_name()} agregado al grupo exitosamente.')
    
    return redirect('dashboard:event_detail', event_id=event_id)


@login_required
@require_http_methods(['POST'])
def event_remove_worker(request, event_id, work_group_id, worker_id):
    event = get_object_or_404(Event, id=event_id)
    work_group = get_object_or_404(EventWorkGroup, id=work_group_id, event=event)
    worker = get_object_or_404(EventWorker, id=worker_id, work_group=work_group)
    
    member_name = worker.member.get_full_name()
    worker.delete()
    
    messages.success(request, f'{member_name} removido del grupo.')
    return redirect('dashboard:event_detail', event_id=event_id)


# Event Budget Views
@login_required
@require_http_methods(['POST'])
def event_update_budget(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    budget, created = EventBudget.objects.get_or_create(event=event)
    
    budget.total_budget = request.POST.get('total_budget', 0)
    budget.target_budget = request.POST.get('target_budget', 0)
    budget.notes = request.POST.get('notes', '')
    budget.save()
    
    messages.success(request, 'Presupuesto actualizado exitosamente.')
    return redirect('dashboard:event_detail', event_id=event_id)


@login_required
@require_http_methods(['POST'])
def event_add_income(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    budget, created = EventBudget.objects.get_or_create(event=event)
    
    income = EventIncome.objects.create(
        budget=budget,
        amount=request.POST.get('amount'),
        source=request.POST.get('source'),
        description=request.POST.get('description'),
        date=request.POST.get('date'),
        receipt_number=request.POST.get('receipt_number', ''),
        notes=request.POST.get('notes', ''),
    )
    
    # Set donor if provided
    donor_id = request.POST.get('donor')
    if donor_id:
        income.donor = Member.objects.get(id=donor_id)
        income.save()
    
    messages.success(request, f'Ingreso de ${income.amount} agregado exitosamente.')
    return redirect('dashboard:event_detail', event_id=event_id)


@login_required
@require_http_methods(['POST'])
def event_remove_income(request, event_id, income_id):
    event = get_object_or_404(Event, id=event_id)
    budget = get_object_or_404(EventBudget, event=event)
    income = get_object_or_404(EventIncome, id=income_id, budget=budget)
    
    amount = income.amount
    income.delete()
    
    messages.success(request, f'Ingreso de ${amount} eliminado.')
    return redirect('dashboard:event_detail', event_id=event_id)


@login_required
@require_http_methods(['POST'])
def event_add_expense(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    budget, created = EventBudget.objects.get_or_create(event=event)
    
    expense = EventExpense.objects.create(
        budget=budget,
        amount=request.POST.get('amount'),
        category=request.POST.get('category'),
        description=request.POST.get('description'),
        vendor=request.POST.get('vendor', ''),
        date=request.POST.get('date'),
        invoice_number=request.POST.get('invoice_number', ''),
        status=request.POST.get('status', 'pending'),
        notes=request.POST.get('notes', ''),
    )
    
    # Set paid_by if provided
    paid_by_id = request.POST.get('paid_by')
    if paid_by_id:
        expense.paid_by = Member.objects.get(id=paid_by_id)
        expense.save()
    
    messages.success(request, f'Egreso de ${expense.amount} agregado exitosamente.')
    return redirect('dashboard:event_detail', event_id=event_id)


@login_required
@require_http_methods(['POST'])
def event_remove_expense(request, event_id, expense_id):
    event = get_object_or_404(Event, id=event_id)
    budget = get_object_or_404(EventBudget, event=event)
    expense = get_object_or_404(EventExpense, id=expense_id, budget=budget)
    
    amount = expense.amount
    expense.delete()
    
    messages.success(request, f'Egreso de ${amount} eliminado.')
    return redirect('dashboard:event_detail', event_id=event_id)


# ── Hospedaje ─────────────────────────────────────────────────────────────────

@login_required
@require_http_methods(['GET', 'POST'])
def event_lodging(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    lodging, _ = EventLodging.objects.get_or_create(event=event)

    if request.method == 'POST':
        lodging.is_enabled = request.POST.get('is_enabled') == 'on'
        lodging.total_needed = request.POST.get('total_needed') or 0
        lodging.notes = request.POST.get('notes', '')
        lodging.save()
        messages.success(request, 'Configuración de hospedaje actualizada.')
        return redirect('dashboard:event_lodging', event_id=event_id)

    hosts = lodging.hosts.prefetch_related('guests').order_by('created_at')
    members = Member.objects.filter(status='active').order_by('last_name', 'first_name')
    context = {
        'event': event,
        'lodging': lodging,
        'hosts': hosts,
        'members': members,
        'total_capacity': lodging.get_total_capacity(),
        'total_assigned': lodging.get_total_assigned(),
        'available_spots': lodging.get_available_spots(),
        'coverage_percentage': lodging.coverage_percentage(),
    }
    return render(request, 'dashboard/events/lodging.html', context)


@login_required
@require_http_methods(['POST'])
def event_lodging_add_host(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    lodging, _ = EventLodging.objects.get_or_create(event=event)
    member_id = request.POST.get('host_member')
    LodgingHost.objects.create(
        lodging=lodging,
        host_id=member_id if member_id else None,
        host_name=request.POST.get('host_name', ''),
        address=request.POST.get('address'),
        capacity=request.POST.get('capacity', 1),
        notes=request.POST.get('notes', ''),
    )
    messages.success(request, 'Anfitrión agregado exitosamente.')
    return redirect('dashboard:event_lodging', event_id=event_id)


@login_required
@require_http_methods(['POST'])
def event_lodging_remove_host(request, event_id, host_id):
    host = get_object_or_404(LodgingHost, id=host_id, lodging__event_id=event_id)
    host.delete()
    messages.success(request, 'Anfitrión eliminado.')
    return redirect('dashboard:event_lodging', event_id=event_id)


@login_required
@require_http_methods(['POST'])
def event_lodging_add_guest(request, event_id, host_id):
    host = get_object_or_404(LodgingHost, id=host_id, lodging__event_id=event_id)
    adults = int(request.POST.get('adults', 1))
    children = int(request.POST.get('children', 0))
    total = adults + children

    if total > host.available_spots():
        messages.error(request, f'No hay suficiente espacio. Disponible: {host.available_spots()} personas.')
        return redirect('dashboard:event_lodging', event_id=event_id)

    member_id = request.POST.get('representative_member')
    LodgingGuest.objects.create(
        host=host,
        representative_id=member_id if member_id else None,
        representative_name=request.POST.get('representative_name', ''),
        adults=adults,
        children=children,
        notes=request.POST.get('notes', ''),
    )
    messages.success(request, f'Grupo de {total} persona(s) asignado.')
    return redirect('dashboard:event_lodging', event_id=event_id)


@login_required
@require_http_methods(['POST'])
def event_lodging_remove_guest(request, event_id, guest_id):
    guest = get_object_or_404(LodgingGuest, id=guest_id, host__lodging__event_id=event_id)
    host = guest.host
    guest.delete()
    # Recalculate
    host.assigned_count = sum(g.total_people() for g in host.guests.all())
    host.save(update_fields=['assigned_count'])
    messages.success(request, 'Grupo eliminado del hospedaje.')
    return redirect('dashboard:event_lodging', event_id=event_id)


# Members Import/Export Views
@login_required
@require_http_methods(['GET'])
def members_export(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Miembros"
    
    # Define headers
    headers = [
        'ID', 'Nombre', 'Apellido', 'Fecha de Nacimiento', 'Género', 
        'Email', 'Teléfono', 'Dirección', 'Fecha de Ingreso', 'Estado',
        'Bautizado', 'Fecha de Bautismo', 'Lugar de Bautismo', 'Bautizado Por', 'Notas'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get members based on filters (same as list view)
    members = Member.objects.all()
    
    # Apply filters from request
    search_query = request.GET.get('q', '').strip()
    if search_query:
        members = members.filter(
            models.Q(first_name__icontains=search_query) |
            models.Q(last_name__icontains=search_query) |
            models.Q(email__icontains=search_query) |
            models.Q(phone__icontains=search_query)
        )
    
    status_filter = request.GET.get('status')
    if status_filter:
        members = members.filter(status=status_filter)
    
    gender_filter = request.GET.get('gender')
    if gender_filter:
        members = members.filter(gender=gender_filter)
    
    baptism_filter = request.GET.get('baptism')
    if baptism_filter == 'baptized':
        members = members.filter(is_baptized=True)
    elif baptism_filter == 'not_baptized':
        members = members.filter(is_baptized=False)
    
    ministry_filter = request.GET.get('ministry')
    if ministry_filter:
        members = members.filter(
            ministry_memberships__ministry_id=ministry_filter,
            ministry_memberships__end_date__isnull=True
        ).distinct()
    
    # Write data
    for row_num, member in enumerate(members.order_by('last_name', 'first_name'), 2):
        ws.cell(row=row_num, column=1, value=str(member.id))
        ws.cell(row=row_num, column=2, value=member.first_name)
        ws.cell(row=row_num, column=3, value=member.last_name)
        ws.cell(row=row_num, column=4, value=member.date_of_birth.strftime('%Y-%m-%d') if member.date_of_birth else '')
        ws.cell(row=row_num, column=5, value=member.get_gender_display() if member.gender else '')
        ws.cell(row=row_num, column=6, value=member.email)
        ws.cell(row=row_num, column=7, value=member.phone)
        ws.cell(row=row_num, column=8, value=member.address)
        ws.cell(row=row_num, column=9, value=member.join_date.strftime('%Y-%m-%d') if member.join_date else '')
        ws.cell(row=row_num, column=10, value=member.get_status_display())
        ws.cell(row=row_num, column=11, value='Sí' if member.is_baptized else 'No')
        ws.cell(row=row_num, column=12, value=member.baptism_date.strftime('%Y-%m-%d') if member.baptism_date else '')
        ws.cell(row=row_num, column=13, value=member.baptism_place)
        ws.cell(row=row_num, column=14, value=member.baptized_by)
        ws.cell(row=row_num, column=15, value=member.notes)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=miembros_export.xlsx'
    wb.save(response)
    
    return response


@login_required
@require_http_methods(['GET', 'POST'])
def members_import(request):
    if request.method == 'POST':
        from openpyxl import load_workbook
        from datetime import datetime
        
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Por favor selecciona un archivo Excel.')
            return redirect('dashboard:members_import')
        
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'El archivo debe ser un archivo Excel (.xlsx o .xls).')
            return redirect('dashboard:members_import')
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            created_count = 0
            updated_count = 0
            error_count = 0
            errors = []
            
            # Skip header row
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    member_id = row[0]
                    first_name = row[1]
                    last_name = row[2]
                    date_of_birth = row[3]
                    gender = row[4]
                    email = row[5] or ''
                    phone = row[6] or ''
                    address = row[7] or ''
                    join_date = row[8]
                    status = row[9]
                    is_baptized = row[10]
                    baptism_date = row[11]
                    baptism_place = row[12] or ''
                    baptized_by = row[13] or ''
                    notes = row[14] or ''
                    
                    # Validate required fields
                    if not first_name or not last_name:
                        errors.append(f'Fila {row_num}: Nombre y apellido son requeridos')
                        error_count += 1
                        continue
                    
                    # Parse dates
                    if isinstance(date_of_birth, str) and date_of_birth:
                        try:
                            date_of_birth = datetime.strptime(date_of_birth, '%Y-%m-%d').date()
                        except:
                            date_of_birth = None
                    
                    if isinstance(join_date, str) and join_date:
                        try:
                            join_date = datetime.strptime(join_date, '%Y-%m-%d').date()
                        except:
                            join_date = timezone.now().date()
                    elif not join_date:
                        join_date = timezone.now().date()
                    
                    if isinstance(baptism_date, str) and baptism_date:
                        try:
                            baptism_date = datetime.strptime(baptism_date, '%Y-%m-%d').date()
                        except:
                            baptism_date = None
                    
                    # Parse gender
                    gender_map = {
                        'Masculino': 'M',
                        'Femenino': 'F',
                        'Otro': 'O',
                        'M': 'M',
                        'F': 'F',
                        'O': 'O',
                    }
                    gender_code = gender_map.get(gender, '')
                    
                    # Parse status
                    status_map = {
                        'Activo': 'active',
                        'Inactivo': 'inactive',
                        'Visitante': 'visitor',
                        'Transferido': 'transferred',
                        'active': 'active',
                        'inactive': 'inactive',
                        'visitor': 'visitor',
                        'transferred': 'transferred',
                    }
                    status_code = status_map.get(status, 'active')
                    
                    # Parse baptized
                    is_baptized_bool = is_baptized in ['Sí', 'Si', 'Yes', 'True', True, 1, '1']
                    
                    # Check if updating or creating
                    if member_id:
                        try:
                            member = Member.objects.get(id=member_id)
                            # Update existing member
                            member.first_name = first_name
                            member.last_name = last_name
                            member.date_of_birth = date_of_birth
                            member.gender = gender_code
                            member.email = email
                            member.phone = phone
                            member.address = address
                            member.join_date = join_date
                            member.status = status_code
                            member.is_baptized = is_baptized_bool
                            member.baptism_date = baptism_date
                            member.baptism_place = baptism_place
                            member.baptized_by = baptized_by
                            member.notes = notes
                            member.save()
                            updated_count += 1
                        except Member.DoesNotExist:
                            # Create new member if ID doesn't exist
                            Member.objects.create(
                                first_name=first_name,
                                last_name=last_name,
                                date_of_birth=date_of_birth,
                                gender=gender_code,
                                email=email,
                                phone=phone,
                                address=address,
                                join_date=join_date,
                                status=status_code,
                                is_baptized=is_baptized_bool,
                                baptism_date=baptism_date,
                                baptism_place=baptism_place,
                                baptized_by=baptized_by,
                                notes=notes,
                            )
                            created_count += 1
                    else:
                        # Create new member
                        Member.objects.create(
                            first_name=first_name,
                            last_name=last_name,
                            date_of_birth=date_of_birth,
                            gender=gender_code,
                            email=email,
                            phone=phone,
                            address=address,
                            join_date=join_date,
                            status=status_code,
                            is_baptized=is_baptized_bool,
                            baptism_date=baptism_date,
                            baptism_place=baptism_place,
                            baptized_by=baptized_by,
                            notes=notes,
                        )
                        created_count += 1
                
                except Exception as e:
                    errors.append(f'Fila {row_num}: {str(e)}')
                    error_count += 1
            
            # Show results
            if created_count > 0:
                messages.success(request, f'{created_count} miembros creados exitosamente.')
            if updated_count > 0:
                messages.success(request, f'{updated_count} miembros actualizados exitosamente.')
            if error_count > 0:
                messages.warning(request, f'{error_count} filas con errores.')
                for error in errors[:10]:  # Show first 10 errors
                    messages.error(request, error)
            
            return redirect('dashboard:members_list')
        
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            return redirect('dashboard:members_import')
    
    return render(request, 'dashboard/members/import.html')


@login_required
@require_http_methods(['GET'])
def members_download_template(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Miembros"
    
    # Define headers
    headers = [
        'ID', 'Nombre', 'Apellido', 'Fecha de Nacimiento', 'Género', 
        'Email', 'Teléfono', 'Dirección', 'Fecha de Ingreso', 'Estado',
        'Bautizado', 'Fecha de Bautismo', 'Lugar de Bautismo', 'Bautizado Por', 'Notas'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add example row
    example_data = [
        '',  # ID (vacío para nuevos miembros)
        'Juan',
        'Pérez',
        '1990-01-15',
        'Masculino',
        'juan.perez@example.com',
        '555-1234',
        'Calle Principal 123',
        '2024-01-01',
        'Activo',
        'Sí',
        '2024-02-15',
        'Iglesia Central',
        'Pastor López',
        'Miembro activo del ministerio de jóvenes'
    ]
    
    for col_num, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col_num, value=value)
    
    # Add instructions
    ws.cell(row=4, column=1, value='INSTRUCCIONES:')
    ws.cell(row=5, column=1, value='1. Deja el ID vacío para crear nuevos miembros')
    ws.cell(row=6, column=1, value='2. Incluye el ID para actualizar miembros existentes')
    ws.cell(row=7, column=1, value='3. Género: Masculino, Femenino, Otro')
    ws.cell(row=8, column=1, value='4. Estado: Activo, Inactivo, Visitante, Transferido')
    ws.cell(row=9, column=1, value='5. Bautizado: Sí o No')
    ws.cell(row=10, column=1, value='6. Fechas en formato: YYYY-MM-DD (ej: 2024-01-15)')
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_miembros.xlsx'
    wb.save(response)
    
    return response


# Ministries Import/Export Views
@login_required
@require_http_methods(['GET'])
def ministries_export(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ministerios"
    
    headers = ['ID', 'Nombre', 'Descripción', 'Líder (ID)', 'Líder (Nombre)', 'Activo']
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ministries = Ministry.objects.all()
    
    search_query = request.GET.get('q', '').strip()
    if search_query:
        ministries = ministries.filter(name__icontains=search_query)
    
    active_filter = request.GET.get('active')
    if active_filter == 'true':
        ministries = ministries.filter(active=True)
    elif active_filter == 'false':
        ministries = ministries.filter(active=False)
    
    for row_num, ministry in enumerate(ministries.order_by('name'), 2):
        ws.cell(row=row_num, column=1, value=str(ministry.id))
        ws.cell(row=row_num, column=2, value=ministry.name)
        ws.cell(row=row_num, column=3, value=ministry.description)
        ws.cell(row=row_num, column=4, value=str(ministry.leader.id) if ministry.leader else '')
        ws.cell(row=row_num, column=5, value=ministry.leader.get_full_name() if ministry.leader else '')
        ws.cell(row=row_num, column=6, value='Sí' if ministry.active else 'No')
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=ministerios_export.xlsx'
    wb.save(response)
    
    return response


@login_required
@require_http_methods(['GET', 'POST'])
def ministries_import(request):
    if request.method == 'POST':
        from openpyxl import load_workbook
        
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Por favor selecciona un archivo Excel.')
            return redirect('dashboard:ministries_import')
        
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'El archivo debe ser un archivo Excel (.xlsx o .xls).')
            return redirect('dashboard:ministries_import')
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            created_count = 0
            updated_count = 0
            error_count = 0
            errors = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    ministry_id = row[0]
                    name = row[1]
                    description = row[2] or ''
                    leader_id = row[3]
                    active = row[5]
                    
                    if not name:
                        errors.append(f'Fila {row_num}: Nombre es requerido')
                        error_count += 1
                        continue
                    
                    active_bool = active in ['Sí', 'Si', 'Yes', 'True', True, 1, '1']
                    
                    leader = None
                    if leader_id:
                        try:
                            leader = Member.objects.get(id=leader_id)
                        except Member.DoesNotExist:
                            errors.append(f'Fila {row_num}: Líder con ID {leader_id} no existe')
                    
                    if ministry_id:
                        try:
                            ministry = Ministry.objects.get(id=ministry_id)
                            ministry.name = name
                            ministry.description = description
                            ministry.leader = leader
                            ministry.active = active_bool
                            ministry.save()
                            updated_count += 1
                        except Ministry.DoesNotExist:
                            Ministry.objects.create(
                                name=name,
                                description=description,
                                leader=leader,
                                active=active_bool,
                            )
                            created_count += 1
                    else:
                        Ministry.objects.create(
                            name=name,
                            description=description,
                            leader=leader,
                            active=active_bool,
                        )
                        created_count += 1
                
                except Exception as e:
                    errors.append(f'Fila {row_num}: {str(e)}')
                    error_count += 1
            
            if created_count > 0:
                messages.success(request, f'{created_count} ministerios creados exitosamente.')
            if updated_count > 0:
                messages.success(request, f'{updated_count} ministerios actualizados exitosamente.')
            if error_count > 0:
                messages.warning(request, f'{error_count} filas con errores.')
                for error in errors[:10]:
                    messages.error(request, error)
            
            return redirect('dashboard:ministries_list')
        
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            return redirect('dashboard:ministries_import')
    
    return render(request, 'dashboard/ministries/import.html')


@login_required
@require_http_methods(['GET'])
def ministries_download_template(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Ministerios"
    
    headers = ['ID', 'Nombre', 'Descripción', 'Líder (ID)', 'Líder (Nombre)', 'Activo']
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    example_data = ['', 'Ministerio de Jóvenes', 'Ministerio dedicado a los jóvenes de la iglesia', '', 'Juan Pérez', 'Sí']
    
    for col_num, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col_num, value=value)
    
    ws.cell(row=4, column=1, value='INSTRUCCIONES:')
    ws.cell(row=5, column=1, value='1. Deja el ID vacío para crear nuevos ministerios')
    ws.cell(row=6, column=1, value='2. Incluye el ID para actualizar ministerios existentes')
    ws.cell(row=7, column=1, value='3. Líder (ID): ID del miembro líder (opcional)')
    ws.cell(row=8, column=1, value='4. Activo: Sí o No')
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_ministerios.xlsx'
    wb.save(response)
    
    return response


# Families Import/Export Views
@login_required
@require_http_methods(['GET'])
def families_export(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Familias"
    
    headers = ['ID', 'Nombre de Familia', 'Dirección', 'Contacto Principal (ID)', 'Contacto Principal (Nombre)', 'Notas']
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    families = Family.objects.all()
    
    search_query = request.GET.get('q', '').strip()
    if search_query:
        families = families.filter(family_name__icontains=search_query)
    
    for row_num, family in enumerate(families.order_by('family_name'), 2):
        ws.cell(row=row_num, column=1, value=str(family.id))
        ws.cell(row=row_num, column=2, value=family.family_name)
        ws.cell(row=row_num, column=3, value=family.address)
        ws.cell(row=row_num, column=4, value=str(family.primary_contact.id) if family.primary_contact else '')
        ws.cell(row=row_num, column=5, value=family.primary_contact.get_full_name() if family.primary_contact else '')
        ws.cell(row=row_num, column=6, value=family.notes)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=familias_export.xlsx'
    wb.save(response)
    
    return response


@login_required
@require_http_methods(['GET', 'POST'])
def families_import(request):
    if request.method == 'POST':
        from openpyxl import load_workbook
        
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Por favor selecciona un archivo Excel.')
            return redirect('dashboard:families_import')
        
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'El archivo debe ser un archivo Excel (.xlsx o .xls).')
            return redirect('dashboard:families_import')
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            created_count = 0
            updated_count = 0
            error_count = 0
            errors = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    family_id = row[0]
                    family_name = row[1]
                    address = row[2] or ''
                    primary_contact_id = row[3]
                    notes = row[5] or ''
                    
                    if not family_name:
                        errors.append(f'Fila {row_num}: Nombre de familia es requerido')
                        error_count += 1
                        continue
                    
                    primary_contact = None
                    if primary_contact_id:
                        try:
                            primary_contact = Member.objects.get(id=primary_contact_id)
                        except Member.DoesNotExist:
                            errors.append(f'Fila {row_num}: Contacto principal con ID {primary_contact_id} no existe')
                    
                    if family_id:
                        try:
                            family = Family.objects.get(id=family_id)
                            family.family_name = family_name
                            family.address = address
                            family.primary_contact = primary_contact
                            family.notes = notes
                            family.save()
                            updated_count += 1
                        except Family.DoesNotExist:
                            Family.objects.create(
                                family_name=family_name,
                                address=address,
                                primary_contact=primary_contact,
                                notes=notes,
                            )
                            created_count += 1
                    else:
                        Family.objects.create(
                            family_name=family_name,
                            address=address,
                            primary_contact=primary_contact,
                            notes=notes,
                        )
                        created_count += 1
                
                except Exception as e:
                    errors.append(f'Fila {row_num}: {str(e)}')
                    error_count += 1
            
            if created_count > 0:
                messages.success(request, f'{created_count} familias creadas exitosamente.')
            if updated_count > 0:
                messages.success(request, f'{updated_count} familias actualizadas exitosamente.')
            if error_count > 0:
                messages.warning(request, f'{error_count} filas con errores.')
                for error in errors[:10]:
                    messages.error(request, error)
            
            return redirect('dashboard:families_list')
        
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            return redirect('dashboard:families_import')
    
    return render(request, 'dashboard/families/import.html')


@login_required
@require_http_methods(['GET'])
def families_download_template(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Familias"
    
    headers = ['ID', 'Nombre de Familia', 'Dirección', 'Contacto Principal (ID)', 'Contacto Principal (Nombre)', 'Notas']
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    example_data = ['', 'Familia Pérez', 'Calle Principal 123, Ciudad', '', 'Juan Pérez', 'Familia activa en la iglesia']
    
    for col_num, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col_num, value=value)
    
    ws.cell(row=4, column=1, value='INSTRUCCIONES:')
    ws.cell(row=5, column=1, value='1. Deja el ID vacío para crear nuevas familias')
    ws.cell(row=6, column=1, value='2. Incluye el ID para actualizar familias existentes')
    ws.cell(row=7, column=1, value='3. Contacto Principal (ID): ID del miembro contacto (opcional)')
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_familias.xlsx'
    wb.save(response)
    
    return response


# Events Import/Export Views
@login_required
@require_http_methods(['GET'])
def events_export(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Eventos"
    
    headers = [
        'ID', 'Nombre', 'Descripción', 'Ubicación', 'Fecha Inicio', 'Fecha Fin',
        'Entrada Gratuita', 'Precio Boleto', 'Capacidad Máxima', 'Tipo de Evento',
        'Estado', 'Organizador (ID)', 'Organizador (Nombre)', 'Notas'
    ]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    events = Event.objects.all()
    
    search_query = request.GET.get('q', '').strip()
    if search_query:
        events = events.filter(name__icontains=search_query)
    
    status_filter = request.GET.get('status')
    if status_filter:
        events = events.filter(status=status_filter)
    
    event_type_filter = request.GET.get('event_type')
    if event_type_filter:
        events = events.filter(event_type=event_type_filter)
    
    for row_num, event in enumerate(events.order_by('-start_date'), 2):
        ws.cell(row=row_num, column=1, value=str(event.id))
        ws.cell(row=row_num, column=2, value=event.name)
        ws.cell(row=row_num, column=3, value=event.description)
        ws.cell(row=row_num, column=4, value=event.location)
        ws.cell(row=row_num, column=5, value=event.start_date.strftime('%Y-%m-%d %H:%M') if event.start_date else '')
        ws.cell(row=row_num, column=6, value=event.end_date.strftime('%Y-%m-%d %H:%M') if event.end_date else '')
        ws.cell(row=row_num, column=7, value='Sí' if event.is_free_entry else 'No')
        ws.cell(row=row_num, column=8, value=float(event.ticket_price) if event.ticket_price else '')
        ws.cell(row=row_num, column=9, value=event.max_capacity if event.max_capacity else '')
        ws.cell(row=row_num, column=10, value=event.get_event_type_display())
        ws.cell(row=row_num, column=11, value=event.get_status_display())
        ws.cell(row=row_num, column=12, value=str(event.organizer.id) if event.organizer else '')
        ws.cell(row=row_num, column=13, value=event.organizer.get_full_name() if event.organizer else '')
        ws.cell(row=row_num, column=14, value=event.notes)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=eventos_export.xlsx'
    wb.save(response)
    
    return response


@login_required
@require_http_methods(['GET', 'POST'])
def events_import(request):
    if request.method == 'POST':
        from openpyxl import load_workbook
        from datetime import datetime
        
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Por favor selecciona un archivo Excel.')
            return redirect('dashboard:events_import')
        
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'El archivo debe ser un archivo Excel (.xlsx o .xls).')
            return redirect('dashboard:events_import')
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            created_count = 0
            updated_count = 0
            error_count = 0
            errors = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    event_id = row[0]
                    name = row[1]
                    description = row[2] or ''
                    location = row[3]
                    start_date = row[4]
                    end_date = row[5]
                    is_free_entry = row[6]
                    ticket_price = row[7]
                    max_capacity = row[8]
                    event_type = row[9]
                    status = row[10]
                    organizer_id = row[11]
                    notes = row[13] or ''
                    
                    if not name or not location or not start_date:
                        errors.append(f'Fila {row_num}: Nombre, ubicación y fecha de inicio son requeridos')
                        error_count += 1
                        continue
                    
                    # Parse dates
                    if isinstance(start_date, str):
                        try:
                            start_date = datetime.strptime(start_date, '%Y-%m-%d %H:%M')
                        except:
                            try:
                                start_date = datetime.strptime(start_date, '%Y-%m-%d')
                            except:
                                errors.append(f'Fila {row_num}: Formato de fecha inicio inválido')
                                error_count += 1
                                continue
                    
                    if end_date and isinstance(end_date, str):
                        try:
                            end_date = datetime.strptime(end_date, '%Y-%m-%d %H:%M')
                        except:
                            try:
                                end_date = datetime.strptime(end_date, '%Y-%m-%d')
                            except:
                                end_date = None
                    
                    is_free_entry_bool = is_free_entry in ['Sí', 'Si', 'Yes', 'True', True, 1, '1']
                    
                    # Parse event type
                    event_type_map = {
                        'Servicio': 'service',
                        'Conferencia': 'conference',
                        'Retiro': 'retreat',
                        'Taller': 'workshop',
                        'Social': 'social',
                        'Alcance': 'outreach',
                        'Otro': 'other',
                    }
                    event_type_code = event_type_map.get(event_type, event_type) if event_type else 'service'
                    
                    # Parse status
                    status_map = {
                        'Borrador': 'draft',
                        'Publicado': 'published',
                        'En Curso': 'ongoing',
                        'Completado': 'completed',
                        'Cancelado': 'cancelled',
                    }
                    status_code = status_map.get(status, status) if status else 'draft'
                    
                    organizer = None
                    if organizer_id:
                        try:
                            organizer = Member.objects.get(id=organizer_id)
                        except Member.DoesNotExist:
                            errors.append(f'Fila {row_num}: Organizador con ID {organizer_id} no existe')
                    
                    if event_id:
                        try:
                            event = Event.objects.get(id=event_id)
                            event.name = name
                            event.description = description
                            event.location = location
                            event.start_date = start_date
                            event.end_date = end_date
                            event.is_free_entry = is_free_entry_bool
                            event.ticket_price = ticket_price if ticket_price else None
                            event.max_capacity = max_capacity if max_capacity else None
                            event.event_type = event_type_code
                            event.status = status_code
                            event.organizer = organizer
                            event.notes = notes
                            event.save()
                            updated_count += 1
                        except Event.DoesNotExist:
                            Event.objects.create(
                                name=name,
                                description=description,
                                location=location,
                                start_date=start_date,
                                end_date=end_date,
                                is_free_entry=is_free_entry_bool,
                                ticket_price=ticket_price if ticket_price else None,
                                max_capacity=max_capacity if max_capacity else None,
                                event_type=event_type_code,
                                status=status_code,
                                organizer=organizer,
                                notes=notes,
                            )
                            created_count += 1
                    else:
                        Event.objects.create(
                            name=name,
                            description=description,
                            location=location,
                            start_date=start_date,
                            end_date=end_date,
                            is_free_entry=is_free_entry_bool,
                            ticket_price=ticket_price if ticket_price else None,
                            max_capacity=max_capacity if max_capacity else None,
                            event_type=event_type_code,
                            status=status_code,
                            organizer=organizer,
                            notes=notes,
                        )
                        created_count += 1
                
                except Exception as e:
                    errors.append(f'Fila {row_num}: {str(e)}')
                    error_count += 1
            
            if created_count > 0:
                messages.success(request, f'{created_count} eventos creados exitosamente.')
            if updated_count > 0:
                messages.success(request, f'{updated_count} eventos actualizados exitosamente.')
            if error_count > 0:
                messages.warning(request, f'{error_count} filas con errores.')
                for error in errors[:10]:
                    messages.error(request, error)
            
            return redirect('dashboard:events_list')
        
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            return redirect('dashboard:events_import')
    
    return render(request, 'dashboard/events/import.html')


@login_required
@require_http_methods(['GET'])
def events_download_template(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Eventos"
    
    headers = [
        'ID', 'Nombre', 'Descripción', 'Ubicación', 'Fecha Inicio', 'Fecha Fin',
        'Entrada Gratuita', 'Precio Boleto', 'Capacidad Máxima', 'Tipo de Evento',
        'Estado', 'Organizador (ID)', 'Organizador (Nombre)', 'Notas'
    ]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    example_data = [
        '', 'Conferencia de Jóvenes 2024', 'Conferencia anual para jóvenes',
        'Auditorio Principal', '2024-06-15 18:00', '2024-06-15 21:00',
        'Sí', '', '200', 'Conferencia', 'Publicado', '', 'Juan Pérez',
        'Evento especial para jóvenes'
    ]
    
    for col_num, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col_num, value=value)
    
    ws.cell(row=4, column=1, value='INSTRUCCIONES:')
    ws.cell(row=5, column=1, value='1. Deja el ID vacío para crear nuevos eventos')
    ws.cell(row=6, column=1, value='2. Incluye el ID para actualizar eventos existentes')
    ws.cell(row=7, column=1, value='3. Fechas en formato: YYYY-MM-DD HH:MM (ej: 2024-06-15 18:00)')
    ws.cell(row=8, column=1, value='4. Tipo de Evento: Servicio, Conferencia, Retiro, Taller, Social, Alcance, Otro')
    ws.cell(row=9, column=1, value='5. Estado: Borrador, Publicado, En Curso, Completado, Cancelado')
    ws.cell(row=10, column=1, value='6. Entrada Gratuita: Sí o No')
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_eventos.xlsx'
    wb.save(response)
    
    return response
